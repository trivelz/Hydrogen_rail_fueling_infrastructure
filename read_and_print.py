# Designing an optimized fueling infrastructure for a hydrogen-powered railway system
# Author: Alessio Trivella

import openpyxl # to read excel files
import matplotlib.pyplot as plt
import support_functions as supp

# print 2D list
def print_2D_list(L,name):
    print(), print(name, end = ""), print(" (%d rows)" %len(L))
    for l in range(0,len(L)):
        for i in range (0, len(L[l]) ):
            print(L[l][i], ' ', end='')
        print()
    
# read the main instance parameters
def read_instance_general(instance_file_path):
    workbook = openpyxl.load_workbook(instance_file_path)
    sheet = workbook['General']

    I = sheet['A2'].value # number of yards
    S = sheet['A3'].value # number of stations
    N = sheet['A4'].value # number of trains
    F = sheet['A5'].value # storage capacity
    delta = sheet['A6'].value # cost per-meter pipeline
    K = sheet['A7'].value # number FD types
    rho = sheet['A8'].value # number FD types

    # construct ranges
    Ir = range(0, I)
    Sr = range(0, S)
    Nr = range(0, N)
    Kr = range(0, K)

    G = [] # fueling dispensers capacities
    for k in range(11, 11 + K):
        cell_value = sheet.cell(row=k, column=1).value
        G.append(cell_value)

    Sn = [] # read information about trains
    for row_number in range(2, 2 + N):
        cell_value = sheet.cell(row=row_number, column=5).value-1
        Sn.append(cell_value)

    Ds = [] # read yard-station distance matrix
    for i in range(2, 2 + I):
        row_data = []
        for s in range(8, 8 + S):
            D_is = sheet.cell(row=i, column=s).value
            row_data.append(D_is) # convert km -> m
        Ds.append(row_data)

    Dn = [] # construct D[i][n]
    for i in range(0, I):
        Dn_row = []
        for n in range(0, N):
            Dn_row.append(Ds[i][Sn[n]])
        Dn.append(Dn_row)

    workbook.close()
    return I,S,N,F,delta,K,rho,Ir,Sr,Nr,Kr,G,Sn,Ds,Dn

# read the yard-specific parameters
def read_instance_yard(instance_file_path, I):

    alpha = [] # cost of yard opening
    C = [] # yard capacity
    beta = [] # cost of placing TT in yard i
    gamma = [] # cost of placing FD type k in yard i
    items = [] # number of items (TT areas, FFI lines, SFI locations)
    area_x = [] # latitude TT [yard][shape][point]
    area_y = [] # longitude TT [yard][shape][point]
    line_x = [] # latitude FFI [yard][line][point]
    line_y = [] # longitude FFI [yard][line][point]
    loca_x = [] # latitude SFI [yard][point]
    loca_y = [] # longitude SFI [yard][point]

    workbook = openpyxl.load_workbook(instance_file_path)
    for i in range(1, 1 + I): # select yard

        # read general information (grey cells)
        sheet = workbook[str(i)]
        alpha.append(sheet.cell(row=1, column=1).value)
        C.append(sheet.cell(row=2, column=1).value)
        beta.append(sheet.cell(row=6, column=1).value)
        gamma_row = []
        gamma_row.append(sheet.cell(row=6, column=7).value) #FFI
        gamma_row.append(sheet.cell(row=6, column=12).value) # SFI
        gamma.append(gamma_row)
        items_row = []
        items_row.append(sheet.cell(row=5, column=1).value) 
        items_row.append(sheet.cell(row=5, column=7).value)
        items_row.append(sheet.cell(row=5, column=12).value)
        items.append(items_row)

        # read coordinates for TT area corners
        area_x_i = [] # 2d matrix
        area_y_i = [] # 2d matrix
        row_num = 9
        for a in range (0, items[i-1][0]):
            points = sheet.cell(row=row_num, column=2).value
            area_x_ip = [] # vector
            area_y_ip = [] # vector
            for p in range (0, points):
                area_x_ip.append(sheet.cell(row=row_num, column=4).value)
                area_y_ip.append(sheet.cell(row=row_num, column=5).value)
                row_num += 1
            area_x_i.append(area_x_ip)
            area_y_i.append(area_y_ip)
        area_x.append(area_x_i)
        area_y.append(area_y_i)

        # read coordinates for FFI line corners
        line_x_i = [] # 2d matrix
        line_y_i = [] # 2d matrix
        row_num = 9
        for a in range (0, items[i-1][1]):
            points = 2
            line_x_ip = [] # vector
            line_y_ip = [] # vector
            for p in range (0, points):
                line_x_ip.append(sheet.cell(row=row_num, column=9).value)
                line_y_ip.append(sheet.cell(row=row_num, column=10).value)
                row_num += 1
            line_x_i.append(line_x_ip)
            line_y_i.append(line_y_ip)
        line_x.append(line_x_i)
        line_y.append(line_y_i)

        # read coordinates for SFI points
        points = items[i-1][2]
        loca_x_i = [] # vector
        loca_y_i = [] # vector
        row_num = 9
        for p in range (0, points):
            loca_x_i.append(sheet.cell(row=row_num, column=13).value)
            loca_y_i.append(sheet.cell(row=row_num, column=14).value)
            row_num += 1
        loca_x.append(loca_x_i)
        loca_y.append(loca_y_i)

    workbook.close()
    return alpha,C,beta,gamma,items,area_x,area_y,line_x,line_y,loca_x,loca_y

# print the instance data on file, for debug
def print_instance_data(I,S,N,F,delta,K,rho,Ir,Sr,G,Sn,Ds,Dn, alpha,C,beta,gamma,
    items,A_lati,A_longi,L_lati,L_longi,P_lati,P_longi,instance_check_file_path):

    with open(instance_check_file_path, 'w') as file:

        print("====================", file=file)
        print("INSTANCE INFORMATION", file=file)
        print("====================", file=file)

        # yards
        print("Number of yards (I):", I, file=file)
        print("Yard capacity (C) [trains]:", C, file=file)
        print("Yard opening cost (alpha) [EUR]:", alpha, file=file)

        # stations
        print("\nNumber of stations (S):", S, file=file)
        print("Train movement cost per-km (rho) [EUR]:", rho, file=file)
        print("Distance yard-station (Ds) [km]:", file=file)
        for i in Ir:
            print(Ds[i], file=file)
        print("Distance yard-train (Dn) [km]:", file=file)
        for i in Ir:
            print(Dn[i], file=file)

        # trains
        print("\nNumber of trains (N):", N, file=file)
        print("Train ending station (Sn):", Sn, file=file) 
        Sn2=[]
        for s in Sr:
            Sn2.append(Sn.count(s))
        print("Train ending at each station:", Sn2, file=file)     

        # infrastructure elements
        print("\nTT capacity (F) [trains]:", F, file=file)
        print("Cost per-km pipeline [EUR]: ", delta, file=file)
        print("Number of FD types:", K, file=file)
        print("FD capacities [trains]:", G, file=file)

        # location coordinates and cost
        print("\n====================", file=file)
        print("STORAGE (TT)", file=file)
        print("====================", end="", file=file)
    
        for i in Ir: 
            print("\n\nYard %d: %d areas" %(i, items[i][0]), end="", file=file)
            print("\nStorage cost in this yard [EUR]:", beta[i], end="", file=file)
            for a in range(0,items[i][0]):
                print("\nArea %d / %d points: " %(a,len(A_lati[i][a])), end="", file=file)
                for p in range(0,len(A_lati[i][a])):
                    print("(%.6f, %.6f) " %(A_lati[i][a][p], A_longi[i][a][p]), end="", file=file)

        # location coordinates and cost
        print("\n\n====================", file=file)
        print("FAST DISPENSERS (FFI)", file=file)
        print("=====================", end="", file=file)
        
        for i in Ir: 
            print("\n\nYard %d: %d lines" %(i, items[i][1]), end="", file=file)
            print("\nFFI cost in this yard [EUR]:", gamma[i][0], end="", file=file)
            for a in range(0,items[i][1]):
                print("\nLine %d: " %a, end="", file=file)
                for p in range(0,len(L_lati[i][a])):
                    print("(%.6f, %.6f) " %(L_lati[i][a][p], L_longi[i][a][p]), end="", file=file)

        # location coordinates and cost
        print("\n\n====================", file=file)
        print("SLOW DISPENSERS (SFI)", file=file)
        print("=====================", file=file)

        for i in Ir: 
            print("\nYard %d: %d locations" %(i, items[i][2]), file=file)
            print("SFI cost in this yard [EUR]:", gamma[i][1], file=file)
            for p in range(0,items[i][2]):
                print("Point %d: (%.6f, %.6f)" %(p, P_lati[i][p], P_longi[i][p]), file=file)

# plots the discretized polygon
def plot_discretized_area(polygon, points, axis_lati, axis_longi, i, a):
    lati_i, longi_i = zip(*points)

    size_max = max(axis_lati,axis_longi) # try to make dimensions more realistic
    longi_dim = max(10*axis_longi/size_max,3)
    lati_dim = max(10*axis_lati/size_max,3)
    plt.figure(figsize=(longi_dim, lati_dim))

    # plot the polygon
    lati_polygon, longi_polygon = zip(*polygon.exterior.coords)
    plt.plot(longi_polygon, lati_polygon, label='Polygon', color='blue')

    # Plot the generated points
    plt.scatter(longi_i, lati_i, label='Generated Points', color='red')
    plt.xlabel('Longitude') # Set labels and title
    plt.ylabel('Latitude')
    plt.title('Yard %d / area %d / number of points %d' %(i,a,len(points)))
    plt.legend() # Add legend
    plt.show() # Show the plot

# plots the discretized line
def plot_discretized_line(L_lati, L_longi, points, i, a):

    min_lati = min(L_lati)
    min_longi = min(L_longi)
    max_lati = max(L_lati)
    max_longi = max(L_longi)
    axis_lati = supp.haversine_distance(min_lati,min_longi,max_lati,min_longi) # latitude distance [m]
    axis_longi = supp.haversine_distance(min_lati,min_longi,min_lati,max_longi) # longitude distance [m]

    size_max = max(axis_lati,axis_longi) # try to make dimensions more realistic
    longi_dim = max(10*axis_longi/size_max,3)
    lati_dim = max(10*axis_lati/size_max,3)
    plt.figure(figsize=(longi_dim, lati_dim))

    # plot the polygon
    plt.plot(L_longi, L_lati, label='Polygon', color='blue')

    # Plot the generated points
    lati_i, longi_i = zip(*points)
    plt.scatter(longi_i, lati_i, label='Generated Points', color='red')
    plt.xlabel('Longitude') # Set labels and title
    plt.ylabel('Latitude')
    plt.title('Yard %d / line %d / number of points %d' %(i,a,len(points)))
    plt.legend() # Add legend
    plt.show() # Show the plot

# print on file the solution to an instance (objective, gap, variables)
def print_solution(solution_file_path,x,v,y,z,w,u,Ir,Kr,I,S,Lr,LTTr,LFDr,gap,runtime,
                   alpha,delta,rho,Dn,Nr,L,E,Sn,G):

    objC = [] # components [C1]-[C5], then full objective (i.e., their sum)
    objC.append( sum(alpha[i] * x[i].x for i in Ir) )
    objC.append( 2 * rho * sum(Dn[i][n] * v[i, n].x for i in Ir for n in Nr) )
    objC.append( sum(L[i][l][4] * y[i, l].x for i in Ir for l in LTTr[i]) ) 
    objC.append( sum(L[i][l][4] * y[i, l].x for i in Ir for l in LFDr[i]) )
    objC.append( delta * sum(E[i][l][m] * u[i, l, m].x for i in Ir for l in Lr[i] for m in Lr[i]) ) 
    objY = [] # cost per yard, then phase 1 objective 
    for i in Ir:
        if x[i].x > 0.1:
            objY.append( sum(L[i][l][4] * y[i, l].x for l in Lr[i]) 
                + delta * sum(E[i][l][m] * u[i, l, m].x for l in Lr[i] for m in Lr[i]) )
        else:
            objY.append(0.0)
    objY.append(objC[0] + objC[1])
    obj = sum(objC)

    with open(solution_file_path, 'w') as file:
        print("====================", file=file)
        print("BEST SOLUTION", file=file)
        print("====================\n", file=file)

        # Objective, gap, runtime
        print("Running time [s]: %.1f" % runtime, file=file)
        print("Optimality gap [%%]: %.3f" % (gap*100), file=file)
        print("\nObjective [EUR]: %.1f" % obj, file=file)
        print("  Breakdown per component:" , file=file)
        for i in range(0,5):
            print(f"  [C{i+1}]: %.1f" % objC[i], file=file)
        print("  Breakdown per yard:" , file=file)
        for i in Ir:
            print(f"  [Y{i}]: %.1f" % objY[i], file=file)
        print(f"   Ph1: %.1f" % objY[-1], file=file)
        print(f"   TOT: %.1f" %(sum(objY)), file=file)

        # Yards and train allocation
        yards = []
        for i in Ir:
            if x[i].x > 0.1:
                yards.append(i)
        print("\nChosen yards (start from 0):", yards, file=file)
        trains_per_Y = [0] * I
        trains_per_YS = [[0] * S for _ in range(I)]
        for i in Ir:
            for n in Nr:
                if v[i, n].x > 0.1:
                    trains_per_Y[i] += 1
                    trains_per_YS[i][Sn[n]] += 1
        print("\nTrains allocated to yards:", trains_per_Y, file=file)
        print("  Breakdown per station:", file=file)
        for i in Ir:
            print(f"  Y{i}:", trains_per_YS[i], file=file)

        # Locations (solution split by yard)
        for i in Ir:
            if x[i].x > 0.1:
                print("\n--------------------", file=file)
                print("YARD ", i ,file=file)
                print("--------------------\n", file=file)

                print("TT locations: ", end="", file=file)
                for l in LTTr[i]:
                    if y[i, l].x > 0.1:
                        print(l," ", end="", file=file)
                for k in Kr:
                    print(f"\nFD locations K{k}: ", end="", file=file)
                    for l in LFDr[i]:
                        if L[i][l][3] == G[k] and y[i, l].x > 0.1:
                            print(l," ", end="", file=file)
                print("\n",file=file)
            # flows
            for l in Lr[i]: 
                for m in Lr[i]: 
                    if w[i, l, m].x > 0.1:
                        print(f"flow {l} -> {m} = {w[i, l, m].x:.3f}", file=file)

    # Print optimal values of decision variables
        print("\n\n====================", file=file)
        print("List of variables", file=file)
        print("====================\n", file=file)
        for i in Ir:
            if x[i].x > 0.1:
                print(f"x[{i}] = {x[i].x}", file=file)
        for i in Ir:
            for l in Lr[i]:
                if y[i, l].x > 0.1:
                    print(f"y[{i}, {l}] = {y[i, l].x:.3f} / ",
                                f"z[{i}, {l}] = {z[i, l].x:.3f}", file=file)
        for i in Ir:
            for l in Lr[i]:
                for m in Lr[i]:
                    if u[i, l, m].x > 0.1:
                        print(f"u[{i}, {l}, {m}] = {u[i, l, m].x:.3f} / ",
                                f"w[{i}, {l}, {m}] = {w[i, l, m].x:.3f}", file=file)

# print on file the solution in a Matlab syntax for plot
def print_solution_Matlab(solution_file_Mat,x,v,y,z,w,u,Ir,I,L,Lr,LTTr,LFDr,objective,
                          alpha,delta,rho,Dn,Nr,E):

    # breakdown of objective components
    obj = [] # first element full objective, then its components [C1]-[C5]
    obj.append(objective) 
    obj.append( sum(alpha[i] * x[i].x for i in Ir) )
    obj.append( 2 * rho * sum(Dn[i][n] * v[i, n].x for i in Ir for n in Nr) )
    obj.append( sum(L[i][l][4] * y[i, l].x for i in Ir for l in LTTr[i]) ) 
    obj.append( sum(L[i][l][4] * y[i, l].x for i in Ir for l in LFDr[i]) )
    obj.append( delta * sum(E[i][l][m] * u[i, l, m].x for i in Ir for l in Lr[i] for m in Lr[i]) )

    with open(solution_file_Mat, 'w') as file:

        # Objective
        print("% Objective (total, components) \nobj = [", end="", file=file) 
        for i in range(0,6):
            print(f" %.0f" % obj[i], end="", file=file)
        print("];\n", file=file)
        
        # x variables
        print("% Yard selection \nx = [", end="", file=file)
        for i in Ir:
            print(f" %d" % x[i].x, end="", file=file)
        print("];\n", file=file)

        # v variable
        trains_per_Y = [0] * I # trains per yard
        for i in Ir:
            for n in Nr:
                if v[i, n].x > 0.1:
                    trains_per_Y[i] += 1
        print("% Train allocation \nv = [", end="", file=file)
        for i in Ir:
            print(f" %d" % trains_per_Y[i], end="", file=file)
        print("];\n", file=file)

        # z variable
        print("% Location variables \nz = [", end="", file=file)
        for i in Ir:
            for l in Lr[i]:
                if y[i, l].x > 0.1:
                    print(f" %d %d %.2f" % (i,l,z[i, l].x),  file=file)
        print("];\n", file=file)

        # w variable
        print("% Connection variables \nw = [", end="", file=file)
        for i in Ir:
            for l in Lr[i]:
                for m in Lr[i]:
                    if u[i, l, m].x > 0.1:
                        print(f" %d %d %d %.2f" % (i, l, m, w[i, l, m].x),  file=file)
        print("];\n", file=file)

        # list of locations
        print("% Locations \nL = [", end="", file=file)
        for i in range(len(L)):
            for j in range(len(L[i])):
                print(f" %d " % i, end="", file=file)
                for k in range(len(L[i][j])):
                    if k == 0 or k == 1:
                        print(f"%.10f " % L[i][j][k], end="", file=file)
                    else:
                        print(f"%d " % L[i][j][k], end="", file=file)

                print("", file=file)
        print(" ];\n", file=file)

# print on file the solution to an instance based on decomposed model
def print_solution_HEUR(solution_file_path,x,v,y,z,w,u,Ir,Kr,I,S,Lr,LTTr,LFDr,gap,runtime,
                        alpha,delta,rho,Dn,Nr,E,L,G,Sn):

    Ir_used = [i for i, value in enumerate(x) if value == 1]

    # breakdown of objective components
    objC = [] # components [C1]-[C5], then full objective (i.e., their sum)
    objC.append( sum(alpha[i] * x[i] for i in Ir_used) )
    objC.append( 2 * rho * sum(Dn[i][n] * v[i][n] for i in Ir_used for n in Nr) )
    objC.append( sum(L[i][l][4] * y[i][l] for i in Ir_used for l in LTTr[i]) ) 
    objC.append( sum(L[i][l][4] * y[i][l] for i in Ir_used for l in LFDr[i]) )
    objC.append( delta * sum(E[i][l][m] * u[i][l][m] for i in Ir_used for l in Lr[i] for m in Lr[i]) ) 
    objY = [] # cost per yard, then phase 1 objective 
    for i in Ir:
        if x[i] > 0.1:
            objY.append( sum(L[i][l][4] * y[i][l] for l in LTTr[i]) + sum(L[i][l][4] * y[i][l] for l in LFDr[i])
                + delta * sum(E[i][l][m] * u[i][l][m] for l in Lr[i] for m in Lr[i]) )
        else:
            objY.append(0.0)
    objY.append(objC[0] + objC[1])
    obj = sum(objC)

    with open(solution_file_path, 'w') as file:
        print("====================", file=file)
        print("BEST SOLUTION", file=file)
        print("====================\n", file=file)

        # Objective, gap, runtime
        print("Running time [s]: %.1f" % runtime, file=file)
        print("Optimality gap subproblems [%]: ", end="", file=file)
        for i in Ir:
            if x[i] > 0.1 and isinstance(gap[i], (float)):
                print("%.3f " %(gap[i]*100), end="", file=file)
            else:
                print("nd ", end="", file=file)
        print("\n\nObjective [EUR]: %.1f" % obj, file=file)
        print("  Breakdown per component:" , file=file)
        for i in range(0,5):
            print(f"  [C{i+1}]: %.1f" % objC[i], file=file)
        print("  Breakdown per yard:" , file=file)
        for i in Ir:
            print(f"  [Y{i}]: %.1f" % objY[i], file=file)
        print(f"   Ph1: %.1f" % objY[-1], file=file)
        print(f"   TOT: %.1f" %(sum(objY)), file=file)
        
        # Yards and train allocation
        yards = []
        for i in Ir:
            if x[i] > 0.1:
                yards.append(i)
        print("\nChosen yards (start from 0):", yards, file=file)
        trains_per_Y = [0] * I
        trains_per_YS = [[0] * S for _ in range(I)]
        for i in Ir:
            for n in Nr:
                if v[i][n] > 0.1:
                    trains_per_Y[i] += 1
                    trains_per_YS[i][Sn[n]] += 1
        print("\nTrains allocated to yards:", trains_per_Y, file=file)
        print("  Breakdown per station:", file=file)
        for i in Ir:
            print(f"  Y{i}:", trains_per_YS[i], file=file)

        # Locations (solution split by yard)
        for i in Ir_used:
            if x[i] > 0.1:
                print("\n--------------------", file=file)
                print("YARD ", i ,file=file)
                print("--------------------\n", file=file)

                print("TT locations: ", end="", file=file)
                for l in LTTr[i]:
                    if y[i][l] > 0.1:
                        print(l," ", end="", file=file)
                for k in Kr:
                    print(f"\nFD locations K{k}: ", end="", file=file)
                    for l in LFDr[i]:
                        if L[i][l][3] == G[k] and y[i][l] > 0.1:
                            print(l," ", end="", file=file)
                print("\n",file=file)
            # flows
            for l in Lr[i]: 
                for m in Lr[i]: 
                    if u[i][l][m] > 0.1:
                        print(f"flow {l} -> {m} = {w[i][l][m]:.3f}", file=file)

    # Print optimal values of decision variables
        print("\n\n====================", file=file)
        print("List of variables", file=file)
        print("====================\n", file=file)
        for i in Ir_used:
            if x[i] > 0.1:
                print(f"x[{i}] = {x[i]}", file=file)
        for i in Ir_used:
            for l in Lr[i]:
                if y[i][l] > 0.1:
                    print(f"y[{i}, {l}] = {y[i][l]:.3f} / ",
                                f"z[{i}, {l}] = {z[i][l]:.3f}", file=file)
        for i in Ir_used:
            for l in Lr[i]:
                for m in Lr[i]:
                    if u[i][l][m] > 0.1:
                        print(f"u[{i}, {l}, {m}] = {u[i][l][m]:.3f} / ",
                                f"w[{i}, {l}, {m}] = {w[i][l][m]:.3f}", file=file)

# print on file the solution in a Matlab syntax for plot
def print_solution_Matlab_HEUR(solution_file_Mat,x,v,y,z,w,u,Ir,I,L,Lr,LTTr,LFDr,objective,
                               alpha,delta,rho,Dn,Nr,E):

        Ir_used = [i for i, value in enumerate(x) if value == 1]

        # breakdown of objective components
        obj = [] # first element full objective, then its components [C1]-[C5]
        obj.append(objective) 
        obj.append( sum(alpha[i] * x[i] for i in Ir_used) )
        obj.append( 2 * rho * sum(Dn[i][n] * v[i][n] for i in Ir_used for n in Nr) )
        obj.append( sum(L[i][l][4] * y[i][l] for i in Ir_used for l in LTTr[i]) ) 
        obj.append( sum(L[i][l][4] * y[i][l] for i in Ir_used for l in LFDr[i]) )
        obj.append( delta * sum(E[i][l][m] * u[i][l][m] for i in Ir_used for l in Lr[i] for m in Lr[i]) )

        with open(solution_file_Mat, 'w') as file:

            # Objective
            print("% Objective (total, components) \nobj = [", end="", file=file) 
            for i in range(0,6):
                print(f" %.0f" % obj[i], end="", file=file)
            print("];\n", file=file)
            
            # x variables
            print("% Yard selection \nx = [", end="", file=file)
            for i in Ir:
                print(f" %d" % x[i], end="", file=file)
            print("];\n", file=file)

            # v variable
            trains_per_Y = [0] * I # trains per yard
            for i in Ir:
                for n in Nr:
                    if v[i][n] > 0.1:
                        trains_per_Y[i] += 1
            print("% Train allocation \nv = [", end="", file=file)
            for i in Ir:
                print(f" %d" % trains_per_Y[i], end="", file=file)
            print("];\n", file=file)

            # z variable
            print("% Location variables \nz = [", end="", file=file)
            for i in Ir_used:
                for l in Lr[i]:
                    if y[i][l] > 0.1:
                        print(f" %d %d %.2f" % (i,l,z[i][l]),  file=file)
            print("];\n", file=file)

            # w variable
            print("% Connection variables \nw = [", end="", file=file)
            for i in Ir_used:
                for l in Lr[i]:
                    for m in Lr[i]:
                        if u[i][l][m] > 0.1:
                            print(f" %d %d %d %.2f" % (i, l, m, w[i][l][m]),  file=file)
            print("];\n", file=file)

            # list of locations
            print("% Locations \nL = [", end="", file=file)
            for i in range(len(L)):
                for j in range(len(L[i])):
                    print(f" %d " % i, end="", file=file)
                    for k in range(len(L[i][j])):
                        if k == 0 or k == 1:
                            print(f"%.10f " % L[i][j][k], end="", file=file)
                        else:
                            print(f"%d " % L[i][j][k], end="", file=file)

                    print("", file=file)
            print(" ];\n", file=file)